import io
import os.path
from re import Pattern

import regex as re
from pathlib import Path
from typing import List, Tuple, Dict, Callable, Any

from openpyxl.reader.excel import load_workbook


class Constants:
    DAY_IN_WEEK: tuple[str] = ("Понеділок", "Вівторок", "Середа", "Четвер", "П'ятниця")
    NUMBER_LESSON: tuple[str] = ('1-2', '3-4', '5-6', '7-8')

    RE_GROUP = re.compile(r'^\p{L}{2,3}-\d{3}$', re.UNICODE)  # відповідає тексту <КІ>-<182> або <КІБ>-<182>


class SheetBasic:
    def __init__(self, sheet, **kwargs):
        self.sheet = sheet

    def run(self, **kwargs):
        # точка запуску утиліти
        pass


class SheetDetectPositionSection(SheetBasic):
    """Визначає позиції початку секцій"""
    day_in_week: Dict[str, int] = {'row': 1, 'column': 0}
    number_lesson: Dict[str, int] = {'row': 1, 'column': 1}
    time_lesson: Dict[str, int] = {'row': 1, 'column': 2}
    group: Dict[str, int] = {'row': 0, 'column': 3}

    @staticmethod
    def create_position(row: int, column: int):
        return dict(
            row=row,
            column=column
        )


class SheetDetectSizeSection(SheetBasic):
    """Визначає розміри секцій"""
    day_in_week: int = 16
    number_lesson: int = 4
    time_lesson: int = 4
    group: int = 3


class SheetTools:

    @staticmethod
    def convert_to_cols(iter_rows) -> List[List[str | None]]:
        """Реалізація iter_cols для Sheet, у режимі читання. Конвертує rows в cols"""
        convert_cols = [list() for _ in iter_rows[0]]

        for cols in iter_rows:
            for index_rows, rows in enumerate(cols):
                convert_cols[index_rows].append(rows)

        return convert_cols


class SheetDetectSize:
    min_row: int = None
    min_col: int = None
    max_row: int = None
    max_col: int = None

    def __init__(self, sheet):
        self.sheet = sheet

    def detect_size(self) -> None:
        """
        Визначає, фактичні розміри сторінки із розкладом. Початкові та кінцеві координати.
        :return:
        """
        list_to_str: Callable[[Any], tuple[str | None, ...]] = lambda lst: tuple(
            map(lambda x: None if x is None else str(x), lst))

        iter_rows = list(self.sheet.iter_rows(values_only=True))
        iter_cols = SheetTools.convert_to_cols(iter_rows)

        # шукаємо, початкові координати
        for index_column, column in enumerate(iter_cols):
            # шукаємо, яка колонка є початковою
            if self._check_exists_in_list(tuple(list_to_str(column)), Constants.DAY_IN_WEEK):
                self.min_col = index_column + 1

                for index_rows, rows in enumerate(iter_rows):
                    # шукаємо, яка рядок є початковою
                    if self._check_exists_in_list_by_pattern(tuple(list_to_str(rows)), Constants.RE_GROUP):
                        self.min_row = index_rows + 1

        # шукаємо, кінцеві координати
        pattern = re.compile(r'^\S*$')
        for index_column, column in enumerate(iter_cols[::-1]):
            if self._check_exists_in_list_by_pattern(tuple(list_to_str(column)), pattern):
                self.max_col = len(iter_cols) - index_column
                break

        for index_rows, rows in enumerate(iter_rows[::-1]):
            if self._check_exists_in_list_by_pattern(tuple(list_to_str(rows)), pattern):
                self.max_row = len(iter_rows) - index_rows + 1
                break

    @staticmethod
    def _check_exists_in_list(checks: tuple, check_fields: tuple):
        """Перевірка, чи колонка, містить відповідні значення"""
        for check_field in check_fields:
            if check_field not in checks:
                return False
        return True

    @staticmethod
    def _check_exists_in_list_by_pattern(checks: tuple, re_pattern: Pattern):
        """Пошук, чи рядок, підпадає під регулярку"""
        for check in checks:
            if check is None:
                continue
            if re_pattern.match(check):
                return True
        return False

    def get_size(self, detect_force: bool = False) -> Dict[str, int | None]:
        """Обраховані поля, в словнику"""
        if detect_force:
            self.detect_size()

        return dict(min_row=self.min_row, max_row=self.max_row, min_col=self.min_col, max_col=self.max_col)


class SheetCut(SheetBasic):
    data: Tuple[Tuple[str | None]] = ()
    bold_cells: Tuple[Tuple[bool]] = ()
    italic_cells: Tuple[Tuple[bool]] = ()

    @staticmethod
    def convert_string_value(field: Any | None):
        """Перетворює колонку або в текст або в None, якщо колонка заповнена пустотою, або None"""

        if field is None or field.value is None:
            # якщо значення None
            return None

        if not isinstance(field.value, str):
            # якщо значення не рядок, конвертувати в рядок
            return str(field.value)

        if re.match(r'^\s*$', field.value):
            # Перевірка, чи рядок є пустим.
            # Перевіряти, чи рядок, немає потреби, тому, що на попередньому рядку вже була перевірка
            return None

        return field.value

    def run(self, size: dict):
        """
        Розрізає сторінку на кілька складових. Сторінку з даними, із жирними клітинками, курсивні
        :return: None
        """
        data = []
        bold_cells = []
        italic_cells = []

        for row in self.sheet.iter_rows(values_only=False, **size):

            data.append(
                tuple(map(self.convert_string_value, row))
            )
            bold_cells_tmp = list()
            italic_cells_tmp = list()

            for cell in row:
                if cell.value is None:
                    bold_cells_tmp.append(False)
                    italic_cells_tmp.append(False)
                    continue

                if cell.font.bold:
                    bold_cells_tmp.append(True)

                if cell.font.italic:
                    italic_cells_tmp.append(True)

            bold_cells.append(tuple(bold_cells_tmp))
            italic_cells.append(tuple(italic_cells_tmp))

        self.data = tuple(data)
        self.bold_cells = tuple(bold_cells)
        self.italic_cells = tuple(italic_cells)


class SheetFieldCut(SheetBasic):
    """Розрізає кожну сторінку, кожну секцію із парою, як один елемент"""
    position_section = SheetDetectPositionSection
    size_section = SheetDetectSizeSection
    sheet_field = dict()

    time_stand_list = dict()

    sheet_lesson_parser = dict()

    @staticmethod
    def slice_two_mas(mas: Tuple[tuple], one_slice: slice, two_slice: slice) -> Tuple[tuple]:
        return tuple(map(lambda x: tuple(x[two_slice]), mas[one_slice]))

    @staticmethod
    def get_day_in_week(iteration: int):
        index = iteration // 4
        return Constants.DAY_IN_WEEK[index]

    @staticmethod
    def get_number_lesson(iteration: int):
        index = iteration % 4
        return Constants.NUMBER_LESSON[index]

    def get_time_stand(self):

        count_lesson = ((len(self.sheet_field['data']) + self.position_section.time_lesson['row']) // 4)
        time_lesson_position = self.position_section.time_lesson
        for index in range(count_lesson):
            time_stand_slice = slice(
                self.size_section.time_lesson * index + time_lesson_position['row'],
                self.size_section.time_lesson * (index + 1) + time_lesson_position['row'],
            )

            day_in_week = Constants.DAY_IN_WEEK[index // 4]
            if self.time_stand_list.get(day_in_week) is None:
                self.time_stand_list[day_in_week] = dict()

            tmp_lesson = list()

            for row in self.sheet['data'][time_stand_slice]:
                tmp_lesson.append(
                    row[time_lesson_position['column']]
                )

            self.time_stand_list[day_in_week][Constants.NUMBER_LESSON[index % 4]] = tmp_lesson.copy()

    def get_name_group(self, iteration: int) -> str:
        size_group_slice_horizontal = slice(
            self.size_section.group + self.size_section.group * iteration,
            self.size_section.group + self.size_section.group * (iteration + 1)
        )

        group_column = self.sheet['data'][self.position_section.group['row']][size_group_slice_horizontal]
        for field in group_column:
            if field is not None and (name_group := Constants.RE_GROUP.match(field)):
                return name_group.string
        return str(iteration)

    def cut_only_field(self):
        """Вирізує таблицю із парами на окремі таблиці"""
        dict_lesson = dict()

        count_groups = len(self.sheet_field['data'][0]) // 3
        for group_index in range(count_groups):

            size_group_slice_horizontal = slice(
                self.size_section.group * group_index,
                self.size_section.group * (group_index + 1)
            )  # ширина колонки

            key_group = self.get_name_group(group_index)

            dict_lesson_in_group = dict()  # словник парами із групами
            for lesson_index in range(len(Constants.DAY_IN_WEEK) * len(Constants.NUMBER_LESSON)):
                size_group_slice_vertical = slice(
                    self.size_section.time_lesson * lesson_index,
                    self.size_section.time_lesson * (lesson_index + 1),
                )

                key_day = self.get_day_in_week(lesson_index)
                if dict_lesson_in_group.get(key_day) is None:
                    dict_lesson_in_group[key_day] = dict()

                number_lesson = self.get_number_lesson(lesson_index)
                dict_lesson_in_group[key_day][number_lesson] = ({
                    'number_lesson': number_lesson,
                    'time_stamp': self.time_stand_list[key_day][number_lesson],
                    'value': dict()
                })
                for key in self.sheet_field.keys():
                    dict_lesson_in_group[key_day][number_lesson]['value'][key] = [
                        item[size_group_slice_horizontal] for item in self.sheet_field[key][
                            size_group_slice_vertical
                        ]
                    ]

                lesson_index += 1

            dict_lesson[key_group] = dict_lesson_in_group.copy()  # копіюємо заповнену групу в словник

        return dict_lesson

    def detect_start_section_with_lesson(self):
        """Вирізує поле із клітинками лише із """
        # початкові координати, звідки починати обрізати сторінки
        one_slice = slice(self.position_section.group['row'] + 1, None)
        two_slice = slice(self.position_section.time_lesson['column'] + 1, None)

        for key in self.sheet.keys():  # проходимося між усіма шарами
            self.sheet_field[key] = self.slice_two_mas(
                self.sheet[key],
                one_slice, two_slice
            )

        return True

    def run(self):
        self.detect_start_section_with_lesson()
        self.get_time_stand()
        self.sheet_lesson_parser = self.cut_only_field()


class ConvertToData(SheetBasic):
    """Конвертує масив значень в дані"""
    sheet: dict

    @staticmethod
    def convert_time_stamp(values: list) -> tuple:
        return tuple([value for value in values if value is not None])

    @staticmethod
    def convert_lesson(values: tuple) -> dict:
        pass

    def find_all_data(self):
        """Проходимося, по всіх елементах"""
        for group, group_value in self.sheet.items():
            group_value: dict
            for day, day_value in group_value.items():
                day_value: dict
                for number_lesson, number_lesson_value in day_value.items():
                    time_stamp_key = 'time_stamp'
                    self.sheet[group][day][number_lesson][time_stamp_key] = self.convert_time_stamp(
                        number_lesson_value[time_stamp_key]
                    )
                    lesson_key = 'value'
                    self.sheet[group][day][number_lesson][lesson_key] = self.convert_lesson(
                        number_lesson_value[lesson_key]
                    )


    def run(self, **kwargs):
        self.find_all_data()


class SchedulerParser:
    byte_stream: io.BytesIO

    sheet: List[Dict[str, Tuple[tuple]]]

    sheet_names: tuple

    def __init__(self, file_bytes: io.BytesIO):
        self.byte_stream = file_bytes
        workbook = load_workbook(self.byte_stream, read_only=True)
        self.sheet_names = tuple(workbook.sheetnames)

        for sheet in self.sheet_names:
            sheet_worker = SheetCut(workbook[sheet])
            sheet_worker.run(SheetDetectSize(workbook[sheet]).get_size(detect_force=True))

            sheet_tmp = dict(
                data=sheet_worker.data,
                italic=sheet_worker.italic_cells,
                bold=sheet_worker.bold_cells,
            )
            if not hasattr(self, 'sheet'):
                self.sheet = []
            self.sheet.append(sheet_tmp)
            sheet_field = SheetFieldCut(sheet_tmp)
            sheet_field.run()
            sheet_lesson_parser_tmp = sheet_field.sheet_lesson_parser
            convert_to_data = ConvertToData(sheet_lesson_parser_tmp)
            convert_to_data.run()
            # заповнення таблиць
        t = print()


if __name__ == '__main__':
    BASE_DIR = Path(__file__).resolve().parent.parent.parent.parent

    path_to_file = os.path.join(BASE_DIR, "Розклад на 1_семестр 2023-2024 н.р.xlsx")
    with open(path_to_file, 'rb') as file:
        scheduler_parser = SchedulerParser(io.BytesIO(file.read()))
        scheduler_parser.detect_size_column(scheduler_parser.sheet_names[0])
